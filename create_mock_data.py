"""
Script to create mock Excel file with the NEW song matching structure
NEW STRUCTURE:
- Row 1: Ausgangssongs (pink)
- Row 2: "Ausgangssong von: Name" in columns B+, "eingegeben von ⬇️" in column A
- Row 3+: Songs with contributors in column A
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def create_mock_excel():
    wb = Workbook()
    wb.remove(wb.active)

    # Mock data with new structure
    worksheets_data = {
        "Runde 4": [
            # Row 1: Ausgangssongs (pink)
            [None, "Peter Schilling - Major Tom", "Ski Aggu - Party Sahne", "The way I are - Timbaland ft Keri Hilson"],
            # Row 2: Ausgangssong contributors
            ["eingegeben von ⬇️", "Ausgangssong von: Phil Phader", "Ausgangssong von: Maksim S.", "Ausgangssong von: Oskar H."],
            # Row 3+: Matching songs with contributors
            ["Alex Awesome (Alexander Bauer)", "Pointer Sisters - I'm so excited", "Miksu - Nachts Wach", "Pink & Redman - Get The Party Started"],
            ["Alex R.", "Live is Life - Opus", "Only 4 Life - Remix Rubi, Farbe Brown", "Sugababes - Push The Button"],
            ["Anusch M.", "Falco - Rock Me Amadeus", "Sido - Medizin (Sonic Empire Remix)", "Lady Gaga - Poker Face"]
        ],
        "Runde 5": [
            [None, "AC/DC - Highway to Hell", "Queen - We Will Rock You", "Guns N' Roses - Sweet Child O' Mine"],
            ["eingegeben von ⬇️", "Ausgangssong von: John D.", "Ausgangssong von: Sarah M.", "Ausgangssong von: Mike K."],
            ["Emma L.", "Led Zeppelin - Whole Lotta Love", "Queen - Another One Bites the Dust", "Bon Jovi - Livin' on a Prayer"],
            ["David R.", "Deep Purple - Smoke on the Water", "The Rolling Stones - Start Me Up", "Journey - Don't Stop Believin'"],
            ["Lisa K.", "Black Sabbath - Paranoid", "Joan Jett - I Love Rock N Roll", "Def Leppard - Pour Some Sugar on Me"]
        ],
        "Runde 6": [
            [None, "Daft Punk - One More Time", "The Prodigy - Firestarter", "Calvin Harris - Summer"],
            ["eingegeben von ⬇️", "Ausgangssong von: Tim S.", "Ausgangssong von: Julia K.", "Ausgangssong von: Marc B."],
            ["Sophie W.", "Modjo - Lady (Hear Me Tonight)", "Chemical Brothers - Block Rockin' Beats", "Avicii - Wake Me Up"],
            ["Felix M.", "Stardust - Music Sounds Better With You", "Fatboy Slim - Right Here Right Now", "David Guetta - Titanium ft Sia"],
            ["Laura H.", "Cassius - Feeling for You", "Basement Jaxx - Where's Your Head At", "Martin Garrix - Animals"]
        ]
    }

    # Create worksheets
    for sheet_name, rows in worksheets_data.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write data row by row
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                if value is not None:
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Apply pink fill to Ausgangssongs (row 1, columns B+)
                    if row_idx == 1 and col_idx > 1:
                        cell.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    wb.save("data/song_matcher_mock.xlsx")
    print("Mock Excel file created with NEW structure: data/song_matcher_mock.xlsx")

if __name__ == "__main__":
    create_mock_excel()
