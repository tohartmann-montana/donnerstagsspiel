"""
Donnerstagsspieler - Find songs and their matching tracks
"""
import os
import streamlit as st
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process
import json
import unicodedata
import re
import html

# Load environment variables from .env file if it exists
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# =============================================================================
# DATABASE MODE FEATURE FLAG
# =============================================================================
# Detect Streamlit Cloud environment (no Excel files available there)
STREAMLIT_CLOUD = (
    os.environ.get("STREAMLIT_SHARING_MODE") is not None or
    os.environ.get("STREAMLIT_SERVER_HEADLESS") == "true"
)

# Set USE_DATABASE=true in environment or .env to use Supabase instead of Excel
# In cloud environment, always use database mode (Excel files not deployed)
USE_DATABASE = os.environ.get("USE_DATABASE", "false").lower() == "true"

# Also check Streamlit secrets for USE_DATABASE flag
try:
    if st.secrets.get("USE_DATABASE", "false").lower() == "true":
        USE_DATABASE = True
except Exception:
    pass

if STREAMLIT_CLOUD:
    USE_DATABASE = True  # Force database mode in cloud

# Import database functions if database mode is enabled
DB_DIAGNOSTICS = None  # Will store diagnostic info for sidebar display

if USE_DATABASE:
    try:
        from db import (
            is_database_available,
            get_database_diagnostics,
            load_all_data_from_db,
            get_all_songs_from_db,
            get_all_contributors_from_db,
            search_songs_db,
            get_song_connections_db,
            get_top_songs_db,
            get_likes_db,
            add_like_db,
            get_contributor_songs_db
        )
        # Get detailed diagnostics instead of just bool
        DB_DIAGNOSTICS = get_database_diagnostics()
        DATABASE_AVAILABLE = DB_DIAGNOSTICS.get('available', False)
    except ImportError as e:
        DATABASE_AVAILABLE = False
        USE_DATABASE = False
        DB_DIAGNOSTICS = {'error': f'Import error: {e}', 'available': False}
    except Exception as e:
        DATABASE_AVAILABLE = False
        DB_DIAGNOSTICS = {'error': f'Startup error: {e}', 'available': False}
else:
    DATABASE_AVAILABLE = False


# ===== Song Normalization =====

def normalize_song_name(song: str) -> str:
    """
    Normalize song name for comparison (not display).
    - Lowercase
    - Remove accents/diacritics
    - Standardize quotes and apostrophes
    - Collapse multiple spaces
    - Normalize dashes/hyphens
    """
    if not song:
        return ""

    # Convert to lowercase
    normalized = song.lower()

    # Remove accents/diacritics (é → e, ü → u, etc.)
    normalized = unicodedata.normalize('NFD', normalized)
    normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')

    # Standardize quotes and apostrophes to straight single quote
    normalized = re.sub(r'[''´`]', "'", normalized)
    normalized = re.sub(r'[""„]', '"', normalized)

    # Normalize various dash characters to standard hyphen
    normalized = re.sub(r'[–—−]', '-', normalized)

    # Collapse multiple spaces/tabs to single space
    normalized = re.sub(r'\s+', ' ', normalized)

    # Normalize spaces around dashes (artist - song format)
    normalized = re.sub(r'\s*-\s*', ' - ', normalized)

    # Strip leading/trailing whitespace
    normalized = normalized.strip()

    return normalized


# ===== Likes Storage Functions =====
LIKES_FILE = Path("data/likes.json")

def load_likes():
    """Load likes from JSON file. Returns empty dict if file doesn't exist."""
    if LIKES_FILE.exists():
        try:
            with open(LIKES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}

def save_likes(likes):
    """Save likes to JSON file."""
    try:
        LIKES_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(LIKES_FILE, 'w', encoding='utf-8') as f:
            json.dump(likes, f, ensure_ascii=False, indent=2)
    except (IOError, OSError) as e:
        st.warning(f"Likes konnten nicht gespeichert werden: {e}")


# ===== Feedback Storage Functions =====
FEEDBACK_FILE = Path("data/feedback.json")

def load_feedback():
    """Load feedback from JSON file."""
    if FEEDBACK_FILE.exists():
        try:
            with open(FEEDBACK_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []

def save_feedback(feedback_list):
    """Save feedback to JSON file."""
    try:
        FEEDBACK_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(FEEDBACK_FILE, 'w', encoding='utf-8') as f:
            json.dump(feedback_list, f, ensure_ascii=False, indent=2)
        return True
    except (IOError, OSError):
        return False

def add_feedback(feedback_type: str, description: str, contact: str = "") -> bool:
    """Add new feedback entry."""
    from datetime import datetime

    feedback_list = load_feedback()
    feedback_list.append({
        "id": len(feedback_list) + 1,
        "type": feedback_type,
        "description": description,
        "contact": contact,
        "theme": "light" if st.session_state.get('light_mode', False) else "dark",
        "timestamp": datetime.now().isoformat()
    })
    return save_feedback(feedback_list)


def add_like(song_name):
    """Add a like to a song and return the new count. Includes debounce protection."""
    import time

    # Debounce: prevent multiple rapid likes (500ms cooldown)
    debounce_key = f"last_like_{hash(song_name) % 10000}"
    current_time = time.time()

    if debounce_key in st.session_state:
        last_like_time = st.session_state[debounce_key]
        if current_time - last_like_time < 0.5:  # 500ms debounce
            # Return current count without incrementing
            if st.session_state.get('using_database', False):
                return get_likes_db().get(song_name, 0)
            else:
                return load_likes().get(song_name, 0)

    st.session_state[debounce_key] = current_time

    # Use database or local storage based on mode
    if st.session_state.get('using_database', False):
        return add_like_db(song_name)
    else:
        likes = load_likes()
        likes[song_name] = likes.get(song_name, 0) + 1
        save_likes(likes)
        return likes[song_name]

def get_like_count(song_name, likes_dict):
    """Get like count for a song from the provided likes dict."""
    return likes_dict.get(song_name, 0)


# ===== Navigation Helpers =====

def navigate_to_song(song_name):
    """Navigate to a song's connections, preserving history for breadcrumbs."""
    # Push current state to history if we're already viewing something
    if st.session_state.selected_song:
        st.session_state.navigation_history.append({
            'type': 'song',
            'value': st.session_state.selected_song
        })
    elif st.session_state.selected_contributor:
        st.session_state.navigation_history.append({
            'type': 'contributor',
            'value': st.session_state.selected_contributor
        })

    # Navigate to new song
    st.session_state.selected_song = song_name
    st.session_state.selected_contributor = None
    st.session_state.page_connections = 1


def navigate_to_contributor(contributor_name):
    """Navigate to a contributor's songs, preserving history for breadcrumbs."""
    # Push current state to history
    if st.session_state.selected_song:
        st.session_state.navigation_history.append({
            'type': 'song',
            'value': st.session_state.selected_song
        })
    elif st.session_state.selected_contributor:
        st.session_state.navigation_history.append({
            'type': 'contributor',
            'value': st.session_state.selected_contributor
        })

    # Navigate to contributor
    st.session_state.selected_contributor = contributor_name
    st.session_state.selected_song = None
    st.session_state.page_contributor = 1


def navigate_back():
    """Navigate back through history, or to search if history empty."""
    if st.session_state.navigation_history:
        # Pop from history
        prev = st.session_state.navigation_history.pop()
        if prev['type'] == 'song':
            st.session_state.selected_song = prev['value']
            st.session_state.selected_contributor = None
        else:
            st.session_state.selected_contributor = prev['value']
            st.session_state.selected_song = None
    else:
        # Go to search, restore previous query
        st.session_state.selected_song = None
        st.session_state.selected_contributor = None
        # Restore last search query via suggestion mechanism
        if st.session_state.last_search_query:
            st.session_state.selected_suggestion = st.session_state.last_search_query


def render_breadcrumbs():
    """Render clickable breadcrumb trail for navigation history."""
    if not st.session_state.navigation_history and not (st.session_state.selected_song or st.session_state.selected_contributor):
        return

    breadcrumb_parts = ["🏠 Suche"]

    # Add history items
    for item in st.session_state.navigation_history:
        if item['type'] == 'song':
            breadcrumb_parts.append(f"🔗 {item['value'][:20]}...")
        else:
            breadcrumb_parts.append(f"👤 {item['value'][:20]}...")

    # Add current item
    if st.session_state.selected_song:
        breadcrumb_parts.append(f"**🔗 {st.session_state.selected_song[:25]}**")
    elif st.session_state.selected_contributor:
        breadcrumb_parts.append(f"**👤 {st.session_state.selected_contributor[:25]}**")

    # Render breadcrumb trail
    st.caption(" → ".join(breadcrumb_parts))


def export_likes_to_csv(likes_dict):
    """Generate CSV content for liked songs."""
    import io
    output = io.StringIO()
    output.write("Song,Likes\n")
    for song, count in sorted(likes_dict.items(), key=lambda x: -x[1]):
        # Escape quotes in song names for CSV
        escaped_song = song.replace('"', '""')
        output.write(f'"{escaped_song}",{count}\n')
    return output.getvalue()


def create_mock_data_if_needed():
    """Create mock Excel file if it doesn't exist"""
    excel_path = Path("data/song_matcher_mock.xlsx")

    if excel_path.exists():
        return

    # Create mock data
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    worksheets_data = {
        "Party Mix": {
            "columns": [
                ["Alex Awesome (Alexander Bauer)", "Alex R.", "Anusch M."],
                [
                    "Peter Schilling - Major Tom",
                    "Pointer Sisters - I'm so excited",
                    "Live is Life - Opus",
                    "Falco - Rock Me Amadeus"
                ],
                [
                    "Ski Aggu - Party Sahne",
                    "Miksu - Nachts Wach",
                    "Only 4 Life - Remix Rubi, Farbe Brown",
                    "Sido - Medizin (Sonic Empire Remix)"
                ],
                [
                    "The way I are - Timbaland ft Keri Hilson",
                    "Pink & Redman - Get The Party Started",
                    "Sugababes - Push The Button",
                    "Lady Gaga - Poker Face"
                ]
            ]
        },
        "Rock Classics": {
            "columns": [
                ["John D.", "Sarah M.", "Mike K."],
                [
                    "AC/DC - Highway to Hell",
                    "Led Zeppelin - Whole Lotta Love",
                    "Deep Purple - Smoke on the Water",
                    "Black Sabbath - Paranoid"
                ],
                [
                    "Queen - We Will Rock You",
                    "Queen - Another One Bites the Dust",
                    "The Rolling Stones - Start Me Up",
                    "Joan Jett - I Love Rock N Roll"
                ],
                [
                    "Guns N' Roses - Sweet Child O' Mine",
                    "Bon Jovi - Livin' on a Prayer",
                    "Journey - Don't Stop Believin'",
                    "Def Leppard - Pour Some Sugar on Me"
                ]
            ]
        },
        "Electronic Vibes": {
            "columns": [
                ["Emma L.", "David R.", "Lisa K."],
                [
                    "Daft Punk - One More Time",
                    "Modjo - Lady (Hear Me Tonight)",
                    "Stardust - Music Sounds Better With You",
                    "Cassius - Feeling for You"
                ],
                [
                    "The Prodigy - Firestarter",
                    "Chemical Brothers - Block Rockin' Beats",
                    "Fatboy Slim - Right Here Right Now",
                    "Basement Jaxx - Where's Your Head At"
                ],
                [
                    "Calvin Harris - Summer",
                    "Avicii - Wake Me Up",
                    "David Guetta - Titanium ft Sia",
                    "Martin Garrix - Animals"
                ]
            ]
        }
    }

    for sheet_name, data in worksheets_data.items():
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, column_data in enumerate(data["columns"], start=1):
            for row_idx, value in enumerate(column_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Pink fill for seed tracks (row 1, columns B onwards)
                if row_idx == 1 and col_idx > 1:
                    cell.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    wb.save(excel_path)

def load_excel_data(file_path):
    """Load all worksheets from Excel file"""
    excel_file = pd.ExcelFile(file_path)
    worksheets = {}

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        worksheets[sheet_name] = df

    return worksheets

@st.cache_data
def get_all_songs(worksheets):
    """Extract all unique songs from all worksheets for autocomplete"""
    all_songs = set()

    for df in worksheets.values():
        for col_idx in range(1, len(df.columns)):
            # Skip empty columns
            if pd.isna(df.iloc[0, col_idx]):
                continue

            # Add seed track from row 1
            seed_track = str(df.iloc[0, col_idx]).strip()
            all_songs.add(seed_track)

            # Add matching songs from row 3+
            column_data = df.iloc[2:, col_idx].dropna()
            column_data = column_data[column_data.astype(str).str.strip() != ''].astype(str)
            all_songs.update(column_data.tolist())

    return sorted(list(all_songs))

@st.cache_data
def get_all_contributors(worksheets):
    """Extract all unique contributors from all worksheets"""
    all_contributors = set()

    for df in worksheets.values():
        # Get contributors from column A (rows 3+)
        contributors = df.iloc[2:, 0].dropna()
        contributors = contributors[contributors.astype(str).str.strip() != ''].astype(str)
        all_contributors.update(contributors.tolist())

        # Get Ausgangssong contributors from row 2
        for col_idx in range(1, len(df.columns)):
            if pd.notna(df.iloc[1, col_idx]):
                text = str(df.iloc[1, col_idx])
                if "Ausgangssong von" in text:
                    contributor = text.replace("Ausgangssong von:", "").replace("Ausgangssong von", "").strip()
                    if contributor:
                        all_contributors.add(contributor)

    return sorted(list(all_contributors))

@st.cache_data
def build_song_index(worksheets):
    """
    Build an optimized index for fast song searching with normalization.
    Returns: dict mapping normalized_song_name -> {variants: set, clusters: list}

    The index groups songs by their normalized name, tracking:
    - All original spelling variants
    - All clusters where any variant appears
    """
    from collections import defaultdict

    # Index structure: normalized_name -> {variants: set(), clusters: list()}
    song_index = defaultdict(lambda: {'variants': set(), 'clusters': []})

    for sheet_name, df in worksheets.items():
        for col_idx in range(1, len(df.columns)):
            if pd.isna(df.iloc[0, col_idx]):
                continue

            # Get seed track and metadata
            seed_track = str(df.iloc[0, col_idx]).strip()

            # Get Ausgangssong contributor
            ausgangssong_contributor_text = str(df.iloc[1, col_idx]) if pd.notna(df.iloc[1, col_idx]) else ""
            if "Ausgangssong von" in ausgangssong_contributor_text:
                seed_contributor = (ausgangssong_contributor_text
                                   .replace("Ausgangssong von:", "")
                                   .replace("Ausgangssong von", "")
                                   .strip())
            else:
                seed_contributor = ausgangssong_contributor_text.strip()

            # Get matching songs
            column_data = df.iloc[2:, col_idx].dropna()
            column_data = column_data[column_data.astype(str).str.strip() != ''].astype(str)
            all_songs = [seed_track] + column_data.tolist()

            # Build contributors dict (keyed by original song name)
            contributors = {seed_track: seed_contributor}
            for idx, song in enumerate(column_data.tolist()):
                actual_row_idx = idx + 2
                contributor = df.iloc[actual_row_idx, 0] if pd.notna(df.iloc[actual_row_idx, 0]) else ""
                contributors[song] = str(contributor).strip()

            # Store cluster metadata
            week_number = col_idx
            round_display = f"{sheet_name}, Woche {week_number}"

            cluster_info = {
                'worksheet': sheet_name,
                'round_display': round_display,
                'seed_track': seed_track,
                'all_songs': all_songs,
                'contributors': contributors,
                'col_idx': col_idx
            }

            # Index all songs by their normalized name
            for song in all_songs:
                normalized = normalize_song_name(song)
                song_index[normalized]['variants'].add(song)
                song_index[normalized]['clusters'].append(cluster_info)

    # Convert sets to sorted lists for JSON serialization (needed for caching)
    result = {}
    for normalized, data in song_index.items():
        result[normalized] = {
            'variants': sorted(list(data['variants'])),
            'clusters': data['clusters'],
            'count': len(data['clusters'])  # Total occurrences
        }

    return result


def get_top_songs(song_index, limit=50):
    """
    Return top songs sorted by occurrence count.
    Filters out non-songs (entries without " - " which indicates Artist - Song format).

    Args:
        song_index: Pre-built song index from build_song_index()
        limit: Maximum number of songs to return (default 50)

    Returns:
        List of dicts with name, normalized, count, variants, clusters
    """
    songs = []
    for normalized, data in song_index.items():
        display_name = data['variants'][0]  # Use first variant as display name

        # Filter: only include entries with " - " (artist-song format)
        # This filters out non-song entries like "Pick the remix", "EX-DUS", etc.
        if " - " not in display_name:
            continue

        songs.append({
            'name': display_name,
            'normalized': normalized,
            'count': data['count'],
            'variants': data['variants'],
            'clusters': data['clusters']  # Include cluster info for drill-down
        })
    return sorted(songs, key=lambda x: x['count'], reverse=True)[:limit]


def get_song_connections(song_name, song_index):
    """
    Find all songs connected to the given song (i.e., songs in the same column/cluster)
    Returns: list of dicts with round_display, seed_track, and connected_songs

    Args:
        song_name: Name of the song to find connections for
        song_index: Pre-built song index from build_song_index()
    """
    # Normalize and lookup
    normalized = normalize_song_name(song_name)

    # In database mode, song_index is None - connections come from DB
    if song_index is None:
        return []

    index_entry = song_index.get(normalized, {'variants': [], 'clusters': []})
    clusters = index_entry['clusters']

    # Transform to expected format (already in correct structure)
    connections = []
    for cluster_info in clusters:
        connections.append({
            'round_display': cluster_info['round_display'],
            'seed_track': cluster_info['seed_track'],
            'all_songs': cluster_info['all_songs'],
            'contributors': cluster_info['contributors']
        })

    return connections


# === PAGINATION HELPERS ===

def render_pagination(items, page_key, items_per_page):
    """
    Calculate pagination state and return paginated items.

    Args:
        items: List of items to paginate
        page_key: Session state key for this pagination context
        items_per_page: Number of items per page

    Returns:
        tuple: (paginated_items, current_page, total_pages, start_num, end_num, total_items)
    """
    if page_key not in st.session_state:
        st.session_state[page_key] = 1

    total_items = len(items)
    total_pages = max(1, (total_items + items_per_page - 1) // items_per_page)

    # Ensure current page is valid
    current_page = min(st.session_state[page_key], total_pages)
    st.session_state[page_key] = current_page

    # Calculate slice indices
    start_idx = (current_page - 1) * items_per_page
    end_idx = min(start_idx + items_per_page, total_items)

    return items[start_idx:end_idx], current_page, total_pages, start_idx + 1, end_idx, total_items


def render_pagination_controls(page_key, current_page, total_pages):
    """
    Render prev/next navigation buttons if more than one page.

    Args:
        page_key: Session state key for this pagination context
        current_page: Current page number
        total_pages: Total number of pages
    """
    if total_pages <= 1:
        return

    col_prev, col_info, col_next = st.columns([1, 2, 1])

    with col_prev:
        if st.button("← Zurück", key=f"{page_key}_prev", disabled=(current_page == 1)):
            st.session_state[page_key] = current_page - 1
            st.rerun()

    with col_info:
        st.markdown(
            f"<div style='text-align:center;padding:0.5rem;color:#94A3B8;'>"
            f"Seite {current_page} von {total_pages}</div>",
            unsafe_allow_html=True
        )

    with col_next:
        if st.button("Weiter →", key=f"{page_key}_next", disabled=(current_page == total_pages)):
            st.session_state[page_key] = current_page + 1
            st.rerun()


def search_songs(query, song_index, fuzzy_threshold=70):
    """
    Search for a song using optimized index and fuzzy matching.
    Returns: list of dicts with worksheet, seed_track, matching_songs, match_scores, and variants

    Args:
        query: Search query string
        song_index: Pre-built song index from build_song_index() with normalized keys
        fuzzy_threshold: Minimum fuzzy match score (0-100). Default 70.
    """
    if not query:
        return []

    # Normalize query for comparison
    query_normalized = normalize_song_name(query)

    # Step 1: Use process.extract() for fast fuzzy matching against normalized keys
    all_normalized_names = list(song_index.keys())

    # Find matches using RapidFuzz's optimized process.extract()
    candidate_matches = process.extract(
        query_normalized,
        all_normalized_names,
        scorer=fuzz.token_sort_ratio,
        limit=100,
        score_cutoff=fuzzy_threshold
    )

    # Step 2: Check for exact substring matches (100% score)
    exact_matches = []
    for normalized_name in all_normalized_names:
        if query_normalized in normalized_name:
            exact_matches.append((normalized_name, 100))

    # Step 3: Combine and deduplicate matches
    all_matches = {}  # normalized_name -> score
    for normalized_name, score, _ in candidate_matches:
        all_matches[normalized_name] = score

    # Add exact matches (override with 100% score)
    for normalized_name, score in exact_matches:
        all_matches[normalized_name] = 100

    if not all_matches:
        return []

    # Step 4: Group matches by cluster (worksheet + column)
    cluster_matches = {}  # (worksheet, col_idx) -> {cluster_info, matched_songs, variants}

    for normalized_name, score in all_matches.items():
        index_entry = song_index[normalized_name]
        variants = index_entry['variants']
        clusters = index_entry['clusters']

        # Use the first variant as the display name (or most common)
        display_name = variants[0] if variants else normalized_name

        for cluster_info in clusters:
            cluster_key = (cluster_info['worksheet'], cluster_info['col_idx'])

            if cluster_key not in cluster_matches:
                cluster_matches[cluster_key] = {
                    'cluster_info': cluster_info,
                    'matched_songs': [],
                    'song_variants': {}  # normalized -> variants list
                }

            # Track the matched song with its display name
            cluster_matches[cluster_key]['matched_songs'].append((display_name, score))
            cluster_matches[cluster_key]['song_variants'][display_name] = variants

    # Step 5: Build results in the expected format
    results = []
    for cluster_key, cluster_data in cluster_matches.items():
        cluster_info = cluster_data['cluster_info']
        matched_songs = cluster_data['matched_songs']
        song_variants = cluster_data['song_variants']

        # Sort matches by score (highest first) and deduplicate
        seen = set()
        unique_matched = []
        for song, score in sorted(matched_songs, key=lambda x: x[1], reverse=True):
            if song not in seen:
                seen.add(song)
                unique_matched.append((song, score))

        results.append({
            'worksheet': cluster_info['worksheet'],
            'round_display': cluster_info['round_display'],
            'seed_track': cluster_info['seed_track'],
            'all_songs': cluster_info['all_songs'],
            'matched_songs': [song for song, score in unique_matched],
            'match_scores': {song: score for song, score in unique_matched},
            'contributors': cluster_info['contributors'],
            'song_variants': song_variants  # Include variant info
        })

    return results

def main():
    st.set_page_config(
        page_title="Donnerstagsspieler",
        page_icon="🎧",
        layout="wide",
        initial_sidebar_state="collapsed"
    )

    # Admin mode disabled for public release
    is_admin = False

    # Initialize theme state
    if 'light_mode' not in st.session_state:
        st.session_state.light_mode = False

    is_light = st.session_state.light_mode

    # Theme colors
    if is_light:
        bg_primary = "#FFFFFF"
        bg_secondary = "#F1F5F9"
        text_primary = "#1E293B"
        text_secondary = "#475569"
        border_color = "#CBD5E1"
        input_bg = "#F8FAFC"
        button_hover_bg = "#E2E8F0"
    else:
        bg_primary = "#0F172A"
        bg_secondary = "#1E293B"
        text_primary = "#E2E8F0"
        text_secondary = "#94A3B8"
        border_color = "#475569"
        input_bg = "#1E293B"
        button_hover_bg = "#2D3E54"

    # Custom CSS for theme
    st.markdown(f"""
        <style>
        /* Import Roboto font from Google Fonts */
        @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');

        /* Base styles - optimized for all devices */
        * {{
            -webkit-tap-highlight-color: rgba(139, 92, 246, 0.3);
        }}

        /* Apply Roboto font globally */
        html, body, [class*="css"] {{
            font-family: 'Roboto', -apple-system, BlinkMacSystemFont, sans-serif !important;
        }}

        /* Override Streamlit's default background */
        .stApp {{
            background-color: {bg_primary} !important;
        }}

        /* Typography hierarchy */
        h1, h2, h3 {{
            font-family: 'Roboto', sans-serif !important;
            font-weight: 700 !important;
            color: {text_primary} !important;
        }}

        /* Section headers */
        .stMarkdown h3 {{
            font-weight: 500 !important;
        }}

        /* Body text */
        p, span, div, label {{
            font-weight: 400;
            color: {text_primary};
        }}

        /* Captions and metadata */
        .stCaption, small, .element-container small {{
            font-weight: 300 !important;
            color: {text_secondary} !important;
        }}

        /* Enlarge search input */
        .stTextInput > div > div > input {{
            font-family: 'Roboto', sans-serif !important;
            font-weight: 400 !important;
            font-size: 1.3rem !important;
            padding: 1rem !important;
            border-radius: 0.5rem !important;
            border: 2px solid #8B5CF6 !important;
            background-color: {input_bg} !important;
            color: {text_primary} !important;
            -webkit-appearance: none !important;
            appearance: none !important;
        }}

        /* Search input focus effect */
        .stTextInput > div > div > input:focus {{
            border-color: #A78BFA !important;
            box-shadow: 0 0 0 3px rgba(139, 92, 246, 0.2) !important;
        }}

        /* Adjust placeholder text */
        .stTextInput > div > div > input::placeholder {{
            color: {text_secondary} !important;
            opacity: 0.7 !important;
        }}

        /* Button styling */
        .stButton > button {{
            font-family: 'Roboto', sans-serif !important;
            font-weight: 500 !important;
            border-radius: 0.5rem !important;
            border: 1px solid {border_color} !important;
            background-color: {bg_secondary} !important;
            color: {text_primary} !important;
            transition: all 0.2s ease !important;
        }}

        .stButton > button:hover {{
            border-color: #8B5CF6 !important;
            background-color: {button_hover_bg} !important;
        }}

        /* Download button styling - force white text */
        .stDownloadButton > button,
        .stDownloadButton > button > div,
        .stDownloadButton > button > div > p,
        .stDownloadButton > button span,
        .stDownloadButton > button p,
        .stDownloadButton button * {{
            font-family: 'Roboto', sans-serif !important;
            font-weight: 500 !important;
            color: #FFFFFF !important;
        }}

        .stDownloadButton > button {{
            border-radius: 0.5rem !important;
            border: 1px solid #8B5CF6 !important;
            background-color: #8B5CF6 !important;
            transition: all 0.2s ease !important;
        }}

        .stDownloadButton > button:hover {{
            background-color: #7C3AED !important;
            border-color: #7C3AED !important;
        }}

        /* Expander styling */
        .streamlit-expanderHeader {{
            font-family: 'Roboto', sans-serif !important;
            font-weight: 500 !important;
            border-radius: 0.5rem !important;
            background-color: {bg_secondary} !important;
            color: {text_primary} !important;
        }}

        /* Expander content */
        .streamlit-expanderContent {{
            background-color: {bg_primary} !important;
            border-color: {border_color} !important;
        }}

        /* Success/Info/Warning boxes */
        .stSuccess, .stInfo, .stWarning {{
            border-radius: 0.5rem !important;
        }}

        /* Tabs styling */
        .stTabs [data-baseweb="tab-list"] {{
            background-color: {bg_secondary} !important;
            border-radius: 0.5rem !important;
        }}

        .stTabs [data-baseweb="tab"] {{
            color: {text_primary} !important;
        }}

        /* Slider styling */
        .stSlider label {{
            color: {text_primary} !important;
        }}

        /* Mobile responsive adjustments */
        @media (max-width: 768px) {{
            /* Smaller search input on mobile */
            .stTextInput > div > div > input {{
                font-size: 1rem !important;
                padding: 0.75rem !important;
            }}

            /* Smaller buttons on mobile */
            .stButton > button {{
                font-size: 0.85rem !important;
                padding: 0.4rem 0.8rem !important;
            }}

            /* Stack columns on mobile */
            .row-widget.stHorizontal {{
                flex-direction: column !important;
            }}

            /* Full width elements on mobile */
            .stTextInput, .stSlider {{
                width: 100% !important;
            }}

            /* Adjust header for mobile */
            .brand-header {{
                flex-direction: column !important;
                align-items: flex-start !important;
                gap: 1rem !important;
                padding: 1rem 0 !important;
            }}

            .brand-divider {{
                width: 100% !important;
                height: 2px !important;
            }}

            .brand-logo {{
                font-size: 2.5rem !important;
            }}

            .brand-title h1 {{
                font-size: 1.5rem !important;
            }}

            .brand-title p {{
                font-size: 0.85rem !important;
            }}

            /* Expander adjustments */
            .streamlit-expanderHeader {{
                font-size: 0.9rem !important;
                padding: 0.5rem !important;
            }}

            /* Better spacing for touch targets */
            .stButton > button {{
                min-height: 44px !important;
                min-width: 44px !important;
            }}
        }}

        /* Tablet adjustments */
        @media (min-width: 769px) and (max-width: 1024px) {{
            .stTextInput > div > div > input {{
                font-size: 1.1rem !important;
            }}

            .brand-logo {{
                font-size: 3rem !important;
            }}
        }}

        /* Custom header branding */
        .brand-header {{
            display: flex;
            align-items: center;
            gap: 1.5rem;
            margin-bottom: 1.5rem;
            padding: 1.5rem 0;
        }}

        .brand-logo {{
            font-size: 3.5rem;
            font-weight: 900;
            line-height: 1;
            letter-spacing: -0.05em;
            color: #8B5CF6;
            font-family: 'Arial Black', sans-serif;
            font-style: italic;
        }}

        .brand-divider {{
            width: 3px;
            height: 50px;
            background: linear-gradient(180deg, #8B5CF6 0%, #A78BFA 100%);
            border-radius: 2px;
        }}

        .brand-title {{
            display: flex;
            flex-direction: column;
        }}

        .brand-title h1 {{
            margin: 0;
            font-size: 2rem;
            color: {text_primary};
            font-family: 'Roboto', sans-serif;
            font-weight: 700;
        }}

        .brand-title p {{
            margin: 0;
            font-size: 0.95rem;
            color: {text_secondary};
            font-family: 'Roboto', sans-serif;
            font-weight: 300;
        }}

        /* Theme toggle styling */
        .theme-toggle {{
            position: absolute;
            right: 1rem;
            top: 0.5rem;
        }}
        </style>
    """, unsafe_allow_html=True)

    # Branded header with theme toggle
    col_header, col_toggle = st.columns([6, 0.5])

    with col_header:
        st.markdown("""
            <div class="brand-header">
                <div class="brand-logo">dus</div>
                <div class="brand-divider"></div>
                <div class="brand-title">
                    <h1>🎧 Donnerstagsspieler</h1>
                    <p>Suche in der Historie des Donnerstagsspiels</p>
                </div>
            </div>
        """, unsafe_allow_html=True)

    with col_toggle:
        # Theme toggle button
        toggle_label = "☀️" if is_light else "🌙"
        toggle_help = "Zum Dunkelmodus wechseln" if is_light else "Zum Hellmodus wechseln"
        if st.button(toggle_label, key="theme_toggle", help=toggle_help):
            st.session_state.light_mode = not st.session_state.light_mode
            st.rerun()

    # Initialize session state for filters
    if 'selected_contributor' not in st.session_state:
        st.session_state.selected_contributor = None
    if 'selected_song' not in st.session_state:
        st.session_state.selected_song = None
    if 'selected_suggestion' not in st.session_state:
        st.session_state.selected_suggestion = None

    # Navigation history for breadcrumb trail (enables infinite drilling)
    if 'navigation_history' not in st.session_state:
        st.session_state.navigation_history = []  # Stack of {'type': 'song'|'contributor', 'value': name}

    # Preserved search context (keeps query when drilling into results)
    if 'last_search_query' not in st.session_state:
        st.session_state.last_search_query = ""

    # Pagination state
    if 'page_search' not in st.session_state:
        st.session_state.page_search = 1
    if 'page_connections' not in st.session_state:
        st.session_state.page_connections = 1
    if 'page_contributor' not in st.session_state:
        st.session_state.page_contributor = 1

    # Track context for pagination reset
    if 'prev_search_query' not in st.session_state:
        st.session_state.prev_search_query = None
    if 'prev_selected_song' not in st.session_state:
        st.session_state.prev_selected_song = None
    if 'prev_selected_contributor' not in st.session_state:
        st.session_state.prev_selected_contributor = None

    # ==========================================================================
    # DATA LOADING - Database or Excel mode
    # ==========================================================================

    # Track data source for conditional logic later
    using_database = USE_DATABASE and DATABASE_AVAILABLE
    st.session_state.using_database = using_database  # Store for use in helper functions

    # ==========================================================================
    # DEBUG SIDEBAR - Show database connection status
    # ==========================================================================
    with st.sidebar:
        st.markdown("### Verbindungsstatus")

        if using_database:
            if DB_DIAGNOSTICS:
                song_count = DB_DIAGNOSTICS.get('song_count', 0)
                st.success(f"Datenbank: {song_count:,} Songs")
            else:
                st.success("Datenbank verbunden")
        else:
            st.warning("Lokaler Modus (Mock-Daten)")

            # Show diagnostic details if database was attempted
            if USE_DATABASE and DB_DIAGNOSTICS:
                st.error("Datenbankverbindung fehlgeschlagen")
                with st.expander("Details anzeigen"):
                    st.write(f"**Secrets Source:** {DB_DIAGNOSTICS.get('secrets_source', 'unknown')}")
                    st.write(f"**Secrets Keys:** {DB_DIAGNOSTICS.get('secrets_keys', [])}")
                    st.write(f"**URL konfiguriert:** {'Ja' if DB_DIAGNOSTICS.get('url_configured') else 'Nein'}")
                    st.write(f"**Key konfiguriert:** {'Ja' if DB_DIAGNOSTICS.get('key_configured') else 'Nein'}")
                    if DB_DIAGNOSTICS.get('url_preview'):
                        st.write(f"**URL:** {DB_DIAGNOSTICS.get('url_preview')}")
                    st.write(f"**Test:** {DB_DIAGNOSTICS.get('connection_test', 'N/A')}")
                    if DB_DIAGNOSTICS.get('error'):
                        st.code(DB_DIAGNOSTICS.get('error'), language=None)
            elif not USE_DATABASE:
                st.info("USE_DATABASE=false")

        st.markdown("---")

    if using_database:
        # === DATABASE MODE ===
        try:
            with st.spinner("Daten werden aus der Datenbank geladen..."):
                # Load data from Supabase
                db_data = load_all_data_from_db()
                if not db_data:
                    st.error("Keine Verbindung zur Datenbank. Bitte Konfiguration prüfen.")
                    return

                total_songs = db_data.get('total_songs', 0)

                # Get autocomplete data
                all_songs = get_all_songs_from_db()
                all_contributors = get_all_contributors_from_db()

                # song_index not needed in DB mode - searches go directly to DB
                song_index = None
                worksheets = None  # Not used in DB mode

                # Load likes from database
                likes = get_likes_db()

            st.info(f"**{total_songs} Songs** in der Datenbank (Cloud-Modus)")

            # Export liked songs to CSV
            if likes:
                csv_data = export_likes_to_csv(likes)
                st.download_button(
                    label=f"📥 Favoriten exportieren ({len(likes)} Songs)",
                    data=csv_data,
                    file_name="donnerstagsspiel_favoriten.csv",
                    mime="text/csv",
                    help="Exportiere alle geliketen Songs als CSV-Datei"
                )
        except Exception as e:
            st.error(f"Fehler beim Laden aus der Datenbank: {e}")
            return
    else:
        # === EXCEL MODE (default) ===
        # Initialize mock data if needed
        create_mock_data_if_needed()

        # Auto-load the database file
        excel_files = list(Path("data").glob("*.xlsx"))

        # Filter out mock data - prefer real data
        real_files = [f for f in excel_files if "mock" not in f.name.lower()]

        if real_files:
            database_file = real_files[0]
        elif excel_files:
            database_file = excel_files[0]
        else:
            st.error("Keine Excel-Dateien im data-Ordner gefunden. Bitte füge eine Excel-Datei hinzu.")
            return

        # Load data with loading indicator
        try:
            with st.spinner("Daten werden geladen..."):
                worksheets = load_excel_data(database_file)

                # Count total songs across all worksheets
                # Row 1 = Ausgangssong, Row 2 = metadata, Row 3+ = matching songs
                total_songs = 0
                for df in worksheets.values():
                    for col_idx in range(1, len(df.columns)):
                        # Count: 1 seed track (row 1) + matching songs (row 3 onwards)
                        seed_count = 1 if pd.notna(df.iloc[0, col_idx]) else 0
                        matching_count = df.iloc[2:, col_idx].dropna().shape[0]
                        total_songs += seed_count + matching_count

                # Extract all songs and contributors for autocomplete
                all_songs = get_all_songs(worksheets)
                all_contributors = get_all_contributors(worksheets)

                # Build optimized song index for fast searching
                song_index = build_song_index(worksheets)

                # Load likes data
                likes = load_likes()

            st.info(f"**{len(worksheets)} Runden** und **{total_songs} Songs** in der Datenbank")

            # Export liked songs to CSV
            if likes:
                csv_data = export_likes_to_csv(likes)
                st.download_button(
                    label=f"📥 Favoriten exportieren ({len(likes)} Songs)",
                    data=csv_data,
                    file_name="donnerstagsspiel_favoriten.csv",
                    mime="text/csv",
                    help="Exportiere alle geliketen Songs als CSV-Datei"
                )
        except Exception as e:
            st.error(f"Fehler beim Laden der Excel-Datei: {e}")
            return

    st.markdown("---")

    # === SCREEN ROUTING ===
    # Show different screens based on session state

    if st.session_state.selected_contributor:
        # === CONTRIBUTOR SCREEN ===
        col_back, col_breadcrumb = st.columns([1, 5])
        with col_back:
            back_label = "← Zurück" if st.session_state.navigation_history else "← Zur Suche"
            if st.button(back_label, key="back_from_contributor"):
                navigate_back()
                st.rerun()
        with col_breadcrumb:
            render_breadcrumbs()

        st.markdown(f"## 👤 Songs von: **{st.session_state.selected_contributor}**")

        # Find all songs by this contributor
        if using_database:
            # Database mode - use database function
            contributor_songs = get_contributor_songs_db(st.session_state.selected_contributor)
        else:
            # Excel mode - iterate through worksheets
            contributor_songs = []
            for sheet_name, df in worksheets.items():
                for col_idx in range(1, len(df.columns)):
                    if pd.isna(df.iloc[0, col_idx]):
                        continue

                    seed_track = str(df.iloc[0, col_idx]).strip()

                    # Check seed track contributor
                    ausgangssong_contributor_text = str(df.iloc[1, col_idx]) if pd.notna(df.iloc[1, col_idx]) else ""
                    if "Ausgangssong von" in ausgangssong_contributor_text:
                        seed_contributor = (ausgangssong_contributor_text
                                           .replace("Ausgangssong von:", "")
                                           .replace("Ausgangssong von", "")
                                           .strip())
                        if seed_contributor == st.session_state.selected_contributor:
                            week_number = col_idx
                            round_display = f"{sheet_name}, Woche {week_number}"
                            contributor_songs.append({
                                'song': seed_track,
                                'round': round_display,
                                'type': '⭐',
                                'type_label': 'Ausgangssong'
                            })

                    # Check other songs
                    column_data = df.iloc[2:, col_idx].dropna()
                    column_data = column_data[column_data.astype(str).str.strip() != ''].astype(str)

                    for idx, song in enumerate(column_data.tolist()):
                        actual_row_idx = idx + 2
                        contributor = df.iloc[actual_row_idx, 0] if pd.notna(df.iloc[actual_row_idx, 0]) else ""
                        if str(contributor).strip() == st.session_state.selected_contributor:
                            week_number = col_idx
                            round_display = f"{sheet_name}, Woche {week_number}"
                            contributor_songs.append({
                                'song': song,
                                'round': round_display,
                                'type': '🎵',
                                'type_label': 'Song'
                            })

        if contributor_songs:
            # Reset page if contributor changed
            if st.session_state.prev_selected_contributor != st.session_state.selected_contributor:
                st.session_state.page_contributor = 1
                st.session_state.prev_selected_contributor = st.session_state.selected_contributor

            # Paginate results (15 per page)
            paginated, page, total_pages, start_num, end_num, total = render_pagination(
                contributor_songs, 'page_contributor', 15
            )

            st.success(f"**{total} Songs** gefunden")
            if total > 15:
                st.caption(f"Zeige {start_num}-{end_num} von {total}")

            for song_idx, song_info in enumerate(paginated):
                col_type, col_song, col_round, col_like = st.columns([0.3, 2.5, 1.5, 0.5])

                with col_type:
                    st.markdown(song_info['type'])

                with col_song:
                    st.markdown(f"**{song_info['song']}**")

                with col_round:
                    st.caption(song_info['round'])

                with col_like:
                    like_count = get_like_count(song_info['song'], likes)
                    like_label = f"❤️ {like_count}" if like_count > 0 else "🤍"
                    # Use start_num + song_idx for unique key across pages
                    if st.button(like_label, key=f"contrib_like_{start_num + song_idx}", help="Song liken"):
                        add_like(song_info['song'])
                        st.rerun()

            # Pagination controls
            render_pagination_controls('page_contributor', page, total_pages)
        else:
            st.info("Keine Songs gefunden")

    elif st.session_state.selected_song:
        # === SONG CONNECTIONS SCREEN ===
        col_back, col_breadcrumb = st.columns([1, 5])
        with col_back:
            back_label = "← Zurück" if st.session_state.navigation_history else "← Zur Suche"
            if st.button(back_label, key="back_from_connections"):
                navigate_back()
                st.rerun()
        with col_breadcrumb:
            render_breadcrumbs()

        st.markdown(f"## 🔗 Verknüpfungen für: **{st.session_state.selected_song}**")

        # Find all connections for this song
        if using_database:
            connections = get_song_connections_db(st.session_state.selected_song)
        else:
            connections = get_song_connections(st.session_state.selected_song, song_index)

        if connections:
            # Reset page if selected song changed
            if st.session_state.prev_selected_song != st.session_state.selected_song:
                st.session_state.page_connections = 1
                st.session_state.prev_selected_song = st.session_state.selected_song

            # Paginate results (10 per page)
            paginated, page, total_pages, start_num, end_num, total = render_pagination(
                connections, 'page_connections', 10
            )

            st.success(f"**{total} Cluster** gefunden")
            if total > 10:
                st.caption(f"Zeige {start_num}-{end_num} von {total}")

            for conn_idx, conn in enumerate(paginated):
                # Use start_num + conn_idx for unique keys across pages
                unique_idx = start_num + conn_idx - 1
                with st.expander(f"📍 {conn['round_display']}", expanded=True):
                    # Show seed track with like button
                    seed_contributor = conn['contributors'].get(conn['seed_track'], "")

                    col_seed, col_seed_like = st.columns([5, 0.5])
                    with col_seed:
                        seed_contributor_display = f" · 👤 {seed_contributor}" if seed_contributor else ""
                        st.markdown(f"⭐ **Ausgangssong:** {conn['seed_track']}{seed_contributor_display}")
                    with col_seed_like:
                        seed_like_count = get_like_count(conn['seed_track'], likes)
                        seed_like_label = f"❤️ {seed_like_count}" if seed_like_count > 0 else "🤍"
                        if st.button(seed_like_label, key=f"conn_seed_like_{unique_idx}", help="Song liken"):
                            add_like(conn['seed_track'])
                            st.rerun()

                    st.markdown("**🎵 Verbundene Songs:**")

                    # Show all songs in this cluster
                    for song_idx, song in enumerate(conn['all_songs']):
                        # Skip the seed track (already shown above)
                        if song == conn['seed_track']:
                            continue

                        contributor = conn['contributors'].get(song, "")

                        col_song, col_drill, col_like = st.columns([4.5, 0.5, 0.5])
                        with col_song:
                            contributor_display = f" · 👤 {contributor}" if contributor else ""
                            # Highlight if this is the selected song
                            if song == st.session_state.selected_song:
                                st.markdown(f"- **{song}** ← *Du bist hier*{contributor_display}")
                            else:
                                st.markdown(f"- {song}{contributor_display}")
                        with col_drill:
                            # Drill-down button to see this song's connections
                            if song != st.session_state.selected_song:
                                if st.button("🔗", key=f"conn_drill_{unique_idx}_{song_idx}", help="Verknüpfungen dieses Songs anzeigen"):
                                    navigate_to_song(song)
                                    st.rerun()
                        with col_like:
                            like_count = get_like_count(song, likes)
                            like_label = f"❤️ {like_count}" if like_count > 0 else "🤍"
                            if st.button(like_label, key=f"conn_like_{unique_idx}_{song_idx}", help="Song liken"):
                                add_like(song)
                                st.rerun()

            # Pagination controls
            render_pagination_controls('page_connections', page, total_pages)
        else:
            st.info("Keine Verknüpfungen gefunden")

    else:
        # === MAIN SCREEN WITH TABS ===
        if is_admin:
            tab_search, tab_bestof, tab_feedback, tab_admin = st.tabs(["🔍 Suche", "🏆 Best Of", "💬 Feedback", "🔐 Admin"])
        else:
            tab_search, tab_bestof, tab_feedback = st.tabs(["🔍 Suche", "🏆 Best Of", "💬 Feedback"])

        with tab_search:
            # === SEARCH TAB ===
            col1, col2 = st.columns([3, 1])

            with col1:
                # Use selected suggestion if available
                default_value = st.session_state.selected_suggestion if st.session_state.selected_suggestion else ""

                search_query = st.text_input(
                    "🔍 Song finden:",
                    value=default_value,
                    placeholder="Songname, Künstler oder Teile davon eingeben...",
                    key="search_input",
                    max_chars=200
                )

                # Clear the suggestion after it's been used
                if st.session_state.selected_suggestion:
                    st.session_state.selected_suggestion = None

                # Show autocomplete suggestions
                if search_query and len(search_query) >= 2:
                    suggestions = [s for s in all_songs if search_query.lower() in s.lower()][:5]
                    if suggestions and search_query not in suggestions:
                        st.caption("💡 Vorschläge:")
                        for idx, suggestion in enumerate(suggestions):
                            # Use index-based key to avoid issues with special characters in song names
                            if st.button(suggestion, key=f"suggest_{idx}", use_container_width=True):
                                st.session_state.selected_suggestion = suggestion
                                st.rerun()

            with col2:
                fuzzy_threshold = st.slider(
                    "Genauigkeit",
                    min_value=50,
                    max_value=100,
                    value=70,
                    step=5,
                    help="Niedriger = toleranter (findet mehr Tippfehler). Höher = strenger."
                )

            if search_query:
                # Use database or local search based on mode
                if using_database:
                    results = search_songs_db(search_query, fuzzy_threshold)
                else:
                    results = search_songs(search_query, song_index, fuzzy_threshold)

                if results:
                    # Reset page if search context changed
                    search_context = f"{search_query}_{fuzzy_threshold}"
                    if st.session_state.prev_search_query != search_context:
                        st.session_state.page_search = 1
                        st.session_state.prev_search_query = search_context

                    # Paginate results (5 per page)
                    paginated, page, total_pages, start_num, end_num, total = render_pagination(
                        results, 'page_search', 5
                    )

                    st.success(f"🎯 {total} Treffer gefunden")
                    if total > 5:
                        st.caption(f"Zeige {start_num}-{end_num} von {total}")

                    for idx, result in enumerate(paginated, start=start_num):
                        with st.expander(f"🎯 Treffer #{idx}: {result['seed_track']} → {result['round_display']}", expanded=True):
                            # Compact header
                            st.markdown(f"📍 **{result['round_display']}**")

                            # Show Ausgangssong with special formatting, like button, and link button
                            seed_contributor = result['contributors'].get(result['seed_track'], "")
                            seed_contributor_display = f" · 👤 **{seed_contributor}**" if seed_contributor else ""

                            col_seed, col_seed_like, col_seed_link = st.columns([5, 0.5, 0.5])
                            with col_seed:
                                st.markdown(f"⭐ **Ausgangssong:** {result['seed_track']}{seed_contributor_display}")
                            with col_seed_like:
                                seed_like_count = get_like_count(result['seed_track'], likes)
                                seed_like_label = f"❤️ {seed_like_count}" if seed_like_count > 0 else "🤍"
                                if st.button(seed_like_label, key=f"like_seed_{idx}", help="Song liken"):
                                    add_like(result['seed_track'])
                                    st.rerun()
                            with col_seed_link:
                                if st.button("🔗", key=f"link_seed_{idx}", help="Verknüpfungen anzeigen"):
                                    # Save search query before navigating
                                    st.session_state.last_search_query = search_query
                                    navigate_to_song(result['seed_track'])
                                    st.rerun()

                            st.markdown("---")
                            st.markdown("**🎵 Passende Songs:**")

                            for song_idx, song in enumerate(result['all_songs'], 1):
                                # Skip the seed track in the list (already shown above)
                                if song == result['seed_track']:
                                    continue

                                contributor = result['contributors'].get(song, "")

                                # Create four columns: song name, like button, link button, contributor button
                                col_song, col_like, col_link, col_contributor = st.columns([2.5, 0.5, 0.5, 1])

                                with col_song:
                                    # Check if this song matched the search
                                    if song in result['matched_songs']:
                                        match_score = result['match_scores'].get(song, 0)

                                        # Color code the badge based on match score
                                        if match_score == 100:
                                            badge_color = "#00CC00"  # Green for perfect match
                                            badge_emoji = "✓"
                                        elif match_score >= 85:
                                            badge_color = "#66CC00"  # Light green
                                            badge_emoji = "~"
                                        elif match_score >= 70:
                                            badge_color = "#FFB800"  # Orange
                                            badge_emoji = "~"
                                        else:
                                            badge_color = "#FF6B6B"  # Red
                                            badge_emoji = "?"

                                        # Check for variants
                                        variants = result.get('song_variants', {}).get(song, [])
                                        variant_badge = ""
                                        if len(variants) > 1:
                                            # Escape variants and join with HTML line break entity
                                            escaped_variants = [html.escape(v, quote=True) for v in variants]
                                            title_text = '&#10;'.join(escaped_variants)
                                            variant_badge = (
                                                f' <span style="background-color: #6366F1; color: white; padding: 2px 8px; '
                                                f'border-radius: 12px; font-size: 0.75em; cursor: help;" '
                                                f'title="{title_text}">'
                                                f'{len(variants)} Varianten</span>'
                                            )

                                        escaped_song = html.escape(song)
                                        st.markdown(
                                            f"{song_idx}. **{escaped_song}** "
                                            f"<span style='background-color: {badge_color}; color: white; padding: 2px 8px; "
                                            f"border-radius: 12px; font-size: 0.8em; font-weight: bold;'>"
                                            f"{badge_emoji} {match_score}%</span>{variant_badge}",
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        # Check for variants even for non-matched songs
                                        normalized = normalize_song_name(song)
                                        # In DB mode, song_index is None
                                        variants = song_index.get(normalized, {}).get('variants', []) if song_index else []
                                        variant_badge = ""
                                        if len(variants) > 1:
                                            # Escape variants and join with HTML line break entity
                                            escaped_variants = [html.escape(v, quote=True) for v in variants]
                                            title_text = '&#10;'.join(escaped_variants)
                                            variant_badge = (
                                                f' <span style="background-color: #6366F1; color: white; padding: 2px 8px; '
                                                f'border-radius: 12px; font-size: 0.75em; cursor: help;" '
                                                f'title="{title_text}">'
                                                f'{len(variants)} Varianten</span>'
                                            )
                                        escaped_song = html.escape(song)
                                        st.markdown(f"{song_idx}. {escaped_song}{variant_badge}", unsafe_allow_html=True)

                                with col_like:
                                    # Like button with heart icon
                                    like_count = get_like_count(song, likes)
                                    like_label = f"❤️ {like_count}" if like_count > 0 else "🤍"
                                    if st.button(like_label, key=f"like_{idx}_{song_idx}", help="Song liken"):
                                        add_like(song)
                                        st.rerun()

                                with col_link:
                                    # Add connection button
                                    if st.button("🔗", key=f"link_{idx}_{song_idx}", help="Verknüpfungen anzeigen"):
                                        # Save search query before navigating
                                        st.session_state.last_search_query = search_query
                                        navigate_to_song(song)
                                        st.rerun()

                                with col_contributor:
                                    if contributor:
                                        # Make contributor clickable
                                        if st.button(f"👤 {contributor}", key=f"contrib_{idx}_{song_idx}", use_container_width=True):
                                            # Save search query before navigating
                                            st.session_state.last_search_query = search_query
                                            navigate_to_contributor(contributor)
                                            st.rerun()

                    # Pagination controls
                    render_pagination_controls('page_search', page, total_pages)
                else:
                    st.warning("Keine Songs gefunden")

        with tab_bestof:
            # === BEST OF TAB ===
            st.markdown("## 🏆 Die beliebtesten Songs")
            st.caption("Songs sortiert nach Anzahl der Auftritte in allen Runden")

            # Get top songs
            if using_database:
                top_songs = get_top_songs_db(limit=50)
            else:
                top_songs = get_top_songs(song_index, limit=50)

            if top_songs:
                # Pagination for Best Of
                if 'page_bestof' not in st.session_state:
                    st.session_state.page_bestof = 1

                paginated, page, total_pages, start_num, end_num, total = render_pagination(
                    top_songs, 'page_bestof', 15
                )

                st.info(f"**Top {total} Songs** nach Häufigkeit")

                for idx, song_data in enumerate(paginated):
                    rank = start_num + idx
                    song_name = song_data['name']
                    variant_count = len(song_data['variants'])
                    variant_info = f" ({variant_count} Varianten)" if variant_count > 1 else ""

                    # Expandable row with song details
                    escaped_song = html.escape(song_name)
                    expander_label = f"{rank}. {escaped_song}{variant_info} — {song_data['count']}x"

                    with st.expander(expander_label, expanded=False):
                        # Show each occurrence with Runde/Woche and Ausgangssong
                        st.markdown("**Vorkommen:**")

                        # clusters data is only available in Excel mode
                        clusters = song_data.get('clusters', [])
                        if clusters:
                            for cluster_idx, cluster in enumerate(clusters):
                                round_display = cluster['round_display']
                                seed_track = cluster['seed_track']

                                col_info, col_link = st.columns([5, 0.5])
                                with col_info:
                                    st.markdown(f"- 📍 **{round_display}** — Ausgangssong: *{html.escape(seed_track)}*")
                                with col_link:
                                    if st.button("🔗", key=f"bestof_cluster_{rank}_{cluster_idx}", help="Verknüpfungen anzeigen"):
                                        navigate_to_song(song_name)
                                        st.rerun()
                        else:
                            # Database mode - show link to view all connections
                            col_info, col_link = st.columns([5, 0.5])
                            with col_info:
                                st.markdown(f"Dieser Song erscheint **{song_data['count']}x** in der Datenbank.")
                            with col_link:
                                if st.button("🔗", key=f"bestof_link_{rank}", help="Alle Verknüpfungen anzeigen"):
                                    navigate_to_song(song_name)
                                    st.rerun()

                # Pagination controls
                render_pagination_controls('page_bestof', page, total_pages)
            else:
                st.info("Keine Songs gefunden")

        with tab_feedback:
            # === FEEDBACK TAB ===
            st.markdown("## 💬 Feedback geben")
            st.markdown("Hast du einen Bug gefunden, eine Idee für ein neues Feature oder sonstiges Feedback? Lass es uns wissen!")

            # Initialize feedback form state
            if 'feedback_submitted' not in st.session_state:
                st.session_state.feedback_submitted = False

            if st.session_state.feedback_submitted:
                st.success("✅ Vielen Dank für dein Feedback! Wir werden es uns ansehen.")
                if st.button("Weiteres Feedback geben"):
                    st.session_state.feedback_submitted = False
                    st.rerun()
            else:
                with st.form("feedback_form"):
                    # Feedback type selection
                    feedback_type = st.selectbox(
                        "Art des Feedbacks",
                        options=["Bug", "Feature", "Other"],
                        format_func=lambda x: {"Bug": "🐛 Bug / Fehler", "Feature": "💡 Feature-Wunsch", "Other": "💬 Sonstiges"}.get(x, x)
                    )

                    # Description
                    feedback_description = st.text_area(
                        "Beschreibung",
                        placeholder="Beschreibe den Bug, das gewünschte Feature oder dein Feedback...",
                        max_chars=2000,
                        height=150
                    )

                    # Optional contact
                    feedback_contact = st.text_input(
                        "Kontakt (optional)",
                        placeholder="E-Mail oder Name, falls wir Rückfragen haben",
                        max_chars=100
                    )

                    # Submit button
                    submitted = st.form_submit_button("📤 Feedback absenden", use_container_width=True)

                    if submitted:
                        if not feedback_description.strip():
                            st.error("Bitte gib eine Beschreibung ein.")
                        else:
                            success = add_feedback(
                                feedback_type=feedback_type,
                                description=feedback_description.strip(),
                                contact=feedback_contact.strip()
                            )
                            if success:
                                st.session_state.feedback_submitted = True
                                st.rerun()
                            else:
                                st.error("Feedback konnte nicht gespeichert werden. Bitte versuche es später erneut.")

        # === ADMIN TAB (only visible with ?admin=true) ===
        if is_admin:
            with tab_admin:
                st.markdown("## 🔐 Feedback-Übersicht")

                feedback_list = load_feedback()

                if feedback_list:
                    st.info(f"**{len(feedback_list)} Feedback-Einträge**")

                    # Export button
                    import io
                    csv_output = io.StringIO()
                    csv_output.write("ID,Typ,Beschreibung,Kontakt,Theme,Timestamp\n")
                    for fb in feedback_list:
                        desc = fb['description'].replace('"', '""').replace('\n', ' ')
                        csv_output.write(f"{fb['id']},{fb['type']},\"{desc}\",{fb.get('contact', '')},{fb.get('theme', '')},{fb['timestamp']}\n")

                    st.download_button(
                        "📥 Feedback als CSV exportieren",
                        csv_output.getvalue(),
                        "feedback_export.csv",
                        "text/csv"
                    )

                    st.markdown("---")

                    # Show feedback entries (newest first)
                    for fb in reversed(feedback_list):
                        type_emoji = {"Bug": "🐛", "Feature": "💡", "Other": "💬"}.get(fb['type'], "📝")

                        with st.expander(f"{type_emoji} #{fb['id']} - {fb['type']} ({fb['timestamp'][:10]})"):
                            st.markdown(f"**Beschreibung:**\n\n{fb['description']}")
                            if fb.get('contact'):
                                st.markdown(f"**Kontakt:** {fb['contact']}")
                            st.caption(f"Theme: {fb.get('theme', 'N/A')} | {fb['timestamp']}")
                else:
                    st.info("Noch kein Feedback eingegangen.")

if __name__ == "__main__":
    main()
